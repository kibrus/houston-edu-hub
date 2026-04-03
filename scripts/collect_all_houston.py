"""
Houston Area College Data Collector
Pulls all institutions within 75mi of Houston (ZIP 77002)
for years 2018-2023 from the College Scorecard API.
Output: data/raw/houston_colleges_all.xlsx
"""

import requests
import os
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

API_KEY     = "B2H5JoCRV54pApGF5FLvirwG3KQesLk8lilyOqts"
BASE_URL    = "https://api.data.gov/ed/collegescorecard/v1/schools"
YEARS       = [2018, 2019, 2020, 2021, 2022, 2023]
HOUSTON_ZIP = "77002"
RADIUS      = "75mi"

# Schools outside the 75mi radius that must be included manually
EXTRA_SCHOOLS = ["University of Houston-Victoria"]

OWNERSHIP_MAP = {1: "Public", 2: "Private Nonprofit", 3: "Private For-Profit"}
DEGREE_MAP    = {0: "Not classified", 1: "Certificate", 2: "Associate's", 3: "Bachelor's", 4: "Graduate"}
SIZE_MAP      = {1: "Small", 2: "Small", 3: "Medium", 4: "Medium", 5: "Large"}
LOCALE_MAP    = {
    11: "City", 12: "City", 13: "City",
    21: "Suburban", 22: "Suburban", 23: "Suburban",
    31: "Town", 32: "Town", 33: "Town",
    41: "Rural", 42: "Rural", 43: "Rural",
}
RELIGIOUS_MAP = {
    22: "American Baptist", 24: "Roman Catholic", 27: "Christian Church",
    28: "Church of Christ", 30: "Episcopal", 33: "Lutheran Church",
    34: "Mennonite", 35: "Methodist", 37: "Pentecostal", 38: "Presbyterian",
    40: "Seventh-day Adventist", 41: "Southern Baptist", 43: "United Methodist",
    44: "Wesleyan", 48: "Interdenominational", 49: "Nondenominational",
    57: "Muslim", 58: "Jewish", -1: "Not Affiliated", 0: "Not Affiliated",
}
MISSION_MAP = {
    "school.minority_serving.historically_black": "Historically Black College and University",
    "school.minority_serving.hispanic":           "Hispanic-Serving Institution",
    "school.minority_serving.aanipi":             "Asian American and Native American Pacific Islander-Serving",
    "school.minority_serving.annh":               "Alaska Native and Native Hawaiian-Serving",
    "school.women_only":                          "Women-Only",
    "school.men_only":                            "Men-Only",
}

STATIC_FIELDS = [
    "id", "school.name", "school.city", "school.state", "school.ownership",
    "school.degrees_awarded.predominant", "school.degrees_awarded.highest",
    "location.lat", "location.lon",
    "school.minority_serving.historically_black",
    "school.minority_serving.hispanic",
    "school.minority_serving.aanipi",
    "school.minority_serving.annh",
    "school.women_only", "school.men_only",
    "school.religious_affiliation", "school.carnegie_size_setting",
    "school.locale", "school.under_investigation",
]

METRIC_FIELD_TEMPLATES = [
    "{year}.cost.tuition.in_state",
    "{year}.cost.tuition.out_of_state",
    "{year}.cost.avg_net_price.overall",
    "{year}.admissions.admission_rate.overall",
    "{year}.completion.rate_suppressed.overall",
    "{year}.student.retention_rate.four_year.full_time",
    "{year}.student.retention_rate.lt_four_year.full_time",
    "{year}.student.size",
    "{year}.student.part_time_share",
    "{year}.student.demographics.student_faculty_ratio",
    "{year}.earnings.10_yrs_after_entry.median",
    "{year}.earnings.10_yrs_after_entry.gt_threshold",
    "{year}.aid.median_debt.completers.overall",
    "{year}.aid.median_debt.completers.monthly_payments",
    "{year}.repayment.3_yr_default_rate",
    "{year}.aid.pell_grant_rate",
    "{year}.aid.federal_loan_rate",
    "{year}.completion.outcome_percentage_suppressed.full_time.first_time.8yr.award_pooled",
    "{year}.completion.outcome_percentage_suppressed.full_time.first_time.8yr.transfer_pooled",
    "{year}.completion.outcome_percentage_suppressed.full_time.first_time.8yr.unknown_pooled",
    "{year}.completion.outcome_percentage_suppressed.full_time.first_time.8yr.still_enrolled_pooled",
    "{year}.admissions.sat_scores.25th_percentile.critical_reading",
    "{year}.admissions.sat_scores.75th_percentile.critical_reading",
    "{year}.admissions.sat_scores.25th_percentile.math",
    "{year}.admissions.sat_scores.75th_percentile.math",
    "{year}.admissions.act_scores.25th_percentile.cumulative",
    "{year}.admissions.act_scores.75th_percentile.cumulative",
]

RACE_FIELD_TEMPLATES = [
    "{year}.student.demographics.race_ethnicity.hispanic",
    "{year}.student.demographics.race_ethnicity.black",
    "{year}.student.demographics.race_ethnicity.white",
    "{year}.student.demographics.race_ethnicity.asian",
    "{year}.student.demographics.race_ethnicity.aian",
    "{year}.student.demographics.race_ethnicity.nhpi",
    "{year}.student.demographics.race_ethnicity.two_or_more",
    "{year}.student.demographics.race_ethnicity.non_resident_alien",
    "{year}.student.demographics.race_ethnicity.unknown",
    "{year}.student.demographics.faculty.race_ethnicity.hispanic",
    "{year}.student.demographics.faculty.race_ethnicity.black",
    "{year}.student.demographics.faculty.race_ethnicity.white",
    "{year}.student.demographics.faculty.race_ethnicity.asian",
    "{year}.student.demographics.faculty.race_ethnicity.aian",
    "{year}.student.demographics.faculty.race_ethnicity.nhpi",
    "{year}.student.demographics.faculty.race_ethnicity.two_or_more",
    "{year}.student.demographics.faculty.race_ethnicity.non_resident_alien",
    "{year}.student.demographics.faculty.race_ethnicity.unknown",
]


def build_fields(year):
    yr = str(year)
    return (
        STATIC_FIELDS
        + [f.replace("{year}", yr) for f in METRIC_FIELD_TEMPLATES]
        + [f.replace("{year}", yr) for f in RACE_FIELD_TEMPLATES]
    )


def fetch_schools(year):
    params = {
        "api_key": API_KEY, "zip": HOUSTON_ZIP, "distance": RADIUS,
        "fields": ",".join(build_fields(year)), "per_page": 100, "page": 0,
    }

    all_results, total_fetched = [], 0
    print(f"Fetching {year}...", end=" ", flush=True)

    while True:
        resp = requests.get(BASE_URL, params=params, timeout=30)
        if resp.status_code != 200:
            print(f"Error {resp.status_code}")
            break
        data    = resp.json()
        results = data.get("results", [])
        total   = data["metadata"]["total"]
        if not results:
            break
        all_results.extend(results)
        total_fetched += len(results)
        if total_fetched >= total:
            break
        params["page"] += 1
        time.sleep(0.3)

    print(f"{total_fetched} schools")
    return all_results


def fetch_school_by_name(name, year):
    params = {
        "api_key":     API_KEY,
        "school.name": name,
        "fields":      ",".join(build_fields(year)),
        "per_page":    1,
    }
    resp    = requests.get(BASE_URL, params=params, timeout=30)
    results = resp.json().get("results", []) if resp.status_code == 200 else []
    if not results:
        print(f"Warning: could not find '{name}' for year {year}")
    return results


def build_table1(all_records_by_year):
    latest_year = max(all_records_by_year.keys())
    rows = []

    for r in all_records_by_year[latest_year]:
        missions = [label for field, label in MISSION_MAP.items() if r.get(field) == 1]
        if not missions:
            specialized = "None"
        elif len(missions) == 1:
            specialized = missions[0]
        else:
            specialized = ", ".join(missions)

        pred    = DEGREE_MAP.get(r.get("school.degrees_awarded.predominant", 0), "")
        highest = DEGREE_MAP.get(r.get("school.degrees_awarded.highest", 0), "")
        awards  = pred if pred == highest else (f"{pred}, {highest}" if pred and highest else pred or highest)

        rows.append({
            "College_ID":            r.get("id"),
            "College_Name":          r.get("school.name"),
            "City":                  r.get("school.city"),
            "State":                 r.get("school.state"),
            "Ownership":             OWNERSHIP_MAP.get(r.get("school.ownership"), "Unknown"),
            "Predominant_Degree":    DEGREE_MAP.get(r.get("school.degrees_awarded.predominant"), "Unknown"),
            "Awards_Offered":        awards,
            "Latitude":              r.get("location.lat"),
            "Longitude":             r.get("location.lon"),
            "Specialized_Mission":   specialized,
            "Religious_Affiliation": RELIGIOUS_MAP.get(r.get("school.religious_affiliation", -1), "Not Affiliated"),
            "WIOA_Programs":         "Yes" if r.get("school.under_investigation") == 1 else "No",
            "Size":                  SIZE_MAP.get(r.get("school.carnegie_size_setting"), "Unknown"),
            "Urbanicity":            LOCALE_MAP.get(r.get("school.locale"), "Unknown"),
        })

    df = pd.DataFrame(rows).drop_duplicates("College_ID").sort_values("College_Name").reset_index(drop=True)
    df["Specialized_Mission"] = df["Specialized_Mission"].fillna("None")
    return df


def build_table2(all_records_by_year):
    rows = []
    for year, records in all_records_by_year.items():
        yr = str(year)
        for r in records:
            pt_share = r.get(f"{yr}.student.part_time_share")
            ft_pct   = round(1 - pt_share, 4) if pt_share is not None else None
            pt_pct   = round(pt_share, 4)      if pt_share is not None else None

            rows.append({
                "College_ID":               r.get("id"),
                "Year":                     year,
                "In_State_Tuition":         r.get(f"{yr}.cost.tuition.in_state"),
                "Out_State_Tuition":        r.get(f"{yr}.cost.tuition.out_of_state"),
                "Net_Annual_Cost":          r.get(f"{yr}.cost.avg_net_price.overall"),
                "Acceptance_Rate":          r.get(f"{yr}.admissions.admission_rate.overall"),
                "Graduation_Rate":          r.get(f"{yr}.completion.rate_suppressed.overall"),
                "Retention_Rate":           (r.get(f"{yr}.student.retention_rate.four_year.full_time")
                                             or r.get(f"{yr}.student.retention_rate.lt_four_year.full_time")),
                "Enrollment":               r.get(f"{yr}.student.size"),
                "FullTime_Pct":             ft_pct,
                "PartTime_Pct":             pt_pct,
                "Student_Faculty_Ratio":    r.get(f"{yr}.student.demographics.student_faculty_ratio"),
                "Median_Earnings":          r.get(f"{yr}.earnings.10_yrs_after_entry.median"),
                "Pct_Earning_More_Than_HS": r.get(f"{yr}.earnings.10_yrs_after_entry.gt_threshold"),
                "Median_Debt":              r.get(f"{yr}.aid.median_debt.completers.overall"),
                "Monthly_Loan_Payment":     r.get(f"{yr}.aid.median_debt.completers.monthly_payments"),
                "Loan_Default_Rate":        r.get(f"{yr}.repayment.3_yr_default_rate"),
                "Pell_Grant_Pct":           r.get(f"{yr}.aid.pell_grant_rate"),
                "Federal_Aid_Pct":          r.get(f"{yr}.aid.federal_loan_rate"),
                "Outcome_Graduated_Pct":    r.get(f"{yr}.completion.outcome_percentage_suppressed.full_time.first_time.8yr.award_pooled"),
                "Outcome_Transferred_Pct":  r.get(f"{yr}.completion.outcome_percentage_suppressed.full_time.first_time.8yr.transfer_pooled"),
                "Outcome_Withdrew_Pct":     r.get(f"{yr}.completion.outcome_percentage_suppressed.full_time.first_time.8yr.unknown_pooled"),
                "Outcome_Enrolled_Pct":     r.get(f"{yr}.completion.outcome_percentage_suppressed.full_time.first_time.8yr.still_enrolled_pooled"),
                "SAT_Reading_25":           r.get(f"{yr}.admissions.sat_scores.25th_percentile.critical_reading"),
                "SAT_Reading_75":           r.get(f"{yr}.admissions.sat_scores.75th_percentile.critical_reading"),
                "SAT_Math_25":              r.get(f"{yr}.admissions.sat_scores.25th_percentile.math"),
                "SAT_Math_75":              r.get(f"{yr}.admissions.sat_scores.75th_percentile.math"),
                "ACT_25":                   r.get(f"{yr}.admissions.act_scores.25th_percentile.cumulative"),
                "ACT_75":                   r.get(f"{yr}.admissions.act_scores.75th_percentile.cumulative"),
            })

    return pd.DataFrame(rows).sort_values(["College_ID", "Year"]).reset_index(drop=True)


def build_table3(all_records_by_year):
    rows = []
    for year, records in all_records_by_year.items():
        yr = str(year)
        for r in records:
            rows.append({
                "College_ID":                   r.get("id"),
                "Year":                         year,
                "Hispanic_Student_Pct":         r.get(f"{yr}.student.demographics.race_ethnicity.hispanic"),
                "Black_Student_Pct":            r.get(f"{yr}.student.demographics.race_ethnicity.black"),
                "White_Student_Pct":            r.get(f"{yr}.student.demographics.race_ethnicity.white"),
                "Asian_Student_Pct":            r.get(f"{yr}.student.demographics.race_ethnicity.asian"),
                "NativeAmerican_Student_Pct":   r.get(f"{yr}.student.demographics.race_ethnicity.aian"),
                "PacificIslander_Student_Pct":  r.get(f"{yr}.student.demographics.race_ethnicity.nhpi"),
                "TwoPlus_Student_Pct":          r.get(f"{yr}.student.demographics.race_ethnicity.two_or_more"),
                "NonResident_Student_Pct":      r.get(f"{yr}.student.demographics.race_ethnicity.non_resident_alien"),
                "Unknown_Student_Pct":          r.get(f"{yr}.student.demographics.race_ethnicity.unknown"),
                "Hispanic_Staff_Pct":           r.get(f"{yr}.student.demographics.faculty.race_ethnicity.hispanic"),
                "Black_Staff_Pct":              r.get(f"{yr}.student.demographics.faculty.race_ethnicity.black"),
                "White_Staff_Pct":              r.get(f"{yr}.student.demographics.faculty.race_ethnicity.white"),
                "Asian_Staff_Pct":              r.get(f"{yr}.student.demographics.faculty.race_ethnicity.asian"),
                "NativeAmerican_Staff_Pct":     r.get(f"{yr}.student.demographics.faculty.race_ethnicity.aian"),
                "PacificIslander_Staff_Pct":    r.get(f"{yr}.student.demographics.faculty.race_ethnicity.nhpi"),
                "TwoPlus_Staff_Pct":            r.get(f"{yr}.student.demographics.faculty.race_ethnicity.two_or_more"),
                "NonResident_Staff_Pct":        r.get(f"{yr}.student.demographics.faculty.race_ethnicity.non_resident_alien"),
                "Unknown_Staff_Pct":            r.get(f"{yr}.student.demographics.faculty.race_ethnicity.unknown"),
            })

    return pd.DataFrame(rows).sort_values(["College_ID", "Year"]).reset_index(drop=True)


def format_sheet(ws, header_color="1F4E79"):
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", start_color=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)
    ws.freeze_panes = "A2"


def save_to_excel(t1, t2, t3, filepath):
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        t1.to_excel(writer, sheet_name="Table1_Colleges",       index=False)
        t2.to_excel(writer, sheet_name="Table2_Metrics_Yearly", index=False)
        t3.to_excel(writer, sheet_name="Table3_Race_Diversity", index=False)
    wb = load_workbook(filepath)
    format_sheet(wb["Table1_Colleges"],       "1F4E79")
    format_sheet(wb["Table2_Metrics_Yearly"], "375623")
    format_sheet(wb["Table3_Race_Diversity"], "843C0C")
    wb.save(filepath)


if __name__ == "__main__":
    print(f"Collecting Houston-area institutions ({YEARS[0]}-{YEARS[-1]})...\n")

    all_records_by_year = {}
    for year in YEARS:
        records      = fetch_schools(year)
        existing_ids = {r.get("id") for r in records}
        for name in EXTRA_SCHOOLS:
            for s in fetch_school_by_name(name, year):
                if s.get("id") not in existing_ids:
                    records.append(s)
                    existing_ids.add(s.get("id"))
            time.sleep(0.3)
        all_records_by_year[year] = records

    print("\nBuilding tables...")
    t1 = build_table1(all_records_by_year)
    t2 = build_table2(all_records_by_year)
    t3 = build_table3(all_records_by_year)

    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_path  = os.path.join(project_root, "data", "raw", "houston_colleges_all.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    save_to_excel(t1, t2, t3, output_path)

    print(f"Done. {t1['College_ID'].nunique()} schools saved to data/raw/houston_colleges_all.xlsx")