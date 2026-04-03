"""
Houston Edu Hub — Curated School Filter
Filters the raw dataset to institutions offering Associate's,
Bachelor's, or Graduate degrees (excludes certificate-only schools).

Input:  data/raw/houston_colleges_all.xlsx
Output: data/curated/houston_colleges_curated.xlsx
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Keep schools where predominant degree is Associate's, Bachelor's, or Graduate
KEEP_DEGREES = ["Associate's", "Bachelor's", "Graduate"]

# Category labels based on ownership and degree
def assign_category(row):
    if row["Ownership"] == "Public" and row["Predominant_Degree"] == "Associate's":
        return "Community College"
    elif row["Ownership"] == "Public":
        return "Public University"
    elif row["Ownership"] == "Private Nonprofit":
        return "Private Nonprofit"
    else:
        return "Private For-Profit"

CATEGORY_COLORS = {
    "Public University":   "DDEEFF",
    "Community College":   "DFF2D8",
    "Private Nonprofit":   "FFF2CC",
    "Private For-Profit":  "FCE4D6",
}


def load_tables(filepath):
    t1 = pd.read_excel(filepath, sheet_name="Table1_Colleges")
    t2 = pd.read_excel(filepath, sheet_name="Table2_Metrics_Yearly")
    t3 = pd.read_excel(filepath, sheet_name="Table3_Race_Diversity")
    return t1, t2, t3


def filter_tables(t1, t2, t3):
    # Filter Table 1 by degree type
    t1_f = t1[t1["Predominant_Degree"].isin(KEEP_DEGREES)].copy()
    t1_f["Category"] = t1_f.apply(assign_category, axis=1)

    # Insert Category right after College_Name
    cols = list(t1_f.columns)
    cols.remove("Category")
    cols.insert(cols.index("College_Name") + 1, "Category")
    t1_f = t1_f[cols]
    t1_f = t1_f.sort_values(["Category", "College_Name"]).reset_index(drop=True)

    # Filter Tables 2 and 3 to match
    curated_ids = set(t1_f["College_ID"].tolist())
    t2_f = t2[t2["College_ID"].isin(curated_ids)].sort_values(["College_ID", "Year"]).reset_index(drop=True)
    t3_f = t3[t3["College_ID"].isin(curated_ids)].sort_values(["College_ID", "Year"]).reset_index(drop=True)

    return t1_f, t2_f, t3_f


def format_sheet(ws, header_color):
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", start_color=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

    ws.freeze_panes = "A2"


def apply_category_colors(wb, t1_f, t2_f, t3_f):
    id_to_cat = dict(zip(t1_f["College_ID"], t1_f["Category"]))

    for sheet_name, df in [
        ("Table1_Colleges",       t1_f),
        ("Table2_Metrics_Yearly", t2_f),
        ("Table3_Race_Diversity", t3_f),
    ]:
        ws     = wb[sheet_name]
        id_col = list(df.columns).index("College_ID")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cid      = row[id_col].value
            category = id_to_cat.get(cid, "")
            color    = CATEGORY_COLORS.get(category, "FFFFFF")
            for cell in row:
                cell.fill      = PatternFill("solid", start_color=color)
                cell.alignment = Alignment(horizontal="center", vertical="center")


def save_curated(t1_f, t2_f, t3_f, filepath):
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        t1_f.to_excel(writer, sheet_name="Table1_Colleges",       index=False)
        t2_f.to_excel(writer, sheet_name="Table2_Metrics_Yearly", index=False)
        t3_f.to_excel(writer, sheet_name="Table3_Race_Diversity", index=False)

    wb = load_workbook(filepath)
    format_sheet(wb["Table1_Colleges"],       "1F4E79")
    format_sheet(wb["Table2_Metrics_Yearly"], "375623")
    format_sheet(wb["Table3_Race_Diversity"], "843C0C")
    apply_category_colors(wb, t1_f, t2_f, t3_f)
    wb.save(filepath)


if __name__ == "__main__":
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_path   = os.path.join(project_root, "data", "raw",     "houston_colleges_all.xlsx")
    output_path  = os.path.join(project_root, "data", "curated", "houston_colleges_curated.xlsx")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    t1, t2, t3       = load_tables(input_path)
    t1_f, t2_f, t3_f = filter_tables(t1, t2, t3)

    print(f"Schools after filtering: {t1_f['College_ID'].nunique()}")
    print()
    for cat in ["Public University", "Community College", "Private Nonprofit", "Private For-Profit"]:
        schools = t1_f[t1_f["Category"] == cat]["College_Name"].tolist()
        if schools:
            print(f"  {cat} ({len(schools)})")
            for s in sorted(schools):
                print(f"    - {s}")
            print()

    save_curated(t1_f, t2_f, t3_f, output_path)
    print(f"Saved to data/curated/houston_colleges_curated.xlsx")